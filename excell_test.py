import os
import datetime

from openpyxl import load_workbook
import subprocess
import shutil
import pandas as pd


def tmp_name(excel_file):
    f = excel_file.split('.')
    return f[0] + "_tmp." + f[1]


def rename_file(excel_file):
    process = subprocess.Popen(["git", "pull"], stdout=subprocess.PIPE)
    output = process.communicate()[0]


def load_excel_wb(excel_file):
    wb2 = load_workbook('test.xlsx')
    print(wb2.sheetnames)
    return wb2


def copy_file_to_tmp(excel_file):
    shutil.copyfile(excel_file, tmp_name(excel_file))


def reset_file_in_git(excel_file):
    process = subprocess.Popen(["git", "checkout", excel_file], stdout=subprocess.PIPE)
    output = process.communicate()[0]


def diff_old_and_new_file(excel_file):
    df1 = pd.read_excel(excel_file)
    df2 = pd.read_excel(tmp_name(excel_file))
    difference = df1[df1 != df2]
    print(difference)
    return difference


def copy_tmp_back(excel_file):
    shutil.copyfile(tmp_name(excel_file), excel_file)


def delete_file(file):
    os.remove(file)


def commit_changes(excel_file):
    process = subprocess.Popen(["git", "add", excel_file], stdout=subprocess.PIPE)
    output = process.communicate()[0]
    process = subprocess.Popen(["git", "commit", "-m", f"File updated {datetime.datetime.now()}"],
                               stdout=subprocess.PIPE)


def main():
    excel_file = "test.xlsx"

    load_excel_wb(excel_file)
    copy_file_to_tmp(excel_file)
    reset_file_in_git(excel_file)

    difference = diff_old_and_new_file(excel_file)

    copy_tmp_back(excel_file)
    delete_file(tmp_name(excel_file))
    commit_changes(excel_file)


if __name__ == "__main__":
    # df1 = pd.read_excel("test.xlsx")
    # df2 = pd.read_excel("test2.xlsx")
    # difference = df1[df1 != df2]
    # print(difference)

    # excel_file = "test.xlsx"
    main()
